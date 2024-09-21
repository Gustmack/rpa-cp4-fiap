import os
import openpyxl
from typing import Dict

# Função para salvar os dados no Excel
def salvar_dados_no_excel(dados: Dict):
    """
    Salva os dados coletados em um arquivo Excel.
    
    :param dados: Dicionário com os dados a serem salvos.
    """
    caminho_pasta = os.path.dirname(os.path.abspath(__file__))
    caminho_arquivo = os.path.join(caminho_pasta, 'dados_ibge_cidade.xlsx')

    wb = openpyxl.load_workbook(caminho_arquivo)
    sheet = wb.active

    # Achar a próxima linha vazia, começando logo após o cabeçalho
    next_row = 2
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Preencher os dados na próxima linha vazia
    formula_pequena_media_grande = f'=IF(E{next_row}>=500000,"GRANDE",IF(AND(E{next_row}>=100000,E{next_row}<500000),"MÉDIA","PEQUENA"))'
    sheet.cell(row=next_row, column=1).value = formula_pequena_media_grande
    sheet.cell(row=next_row, column=2, value=dados["nome_municipio"].upper())
    sheet.cell(row=next_row, column=3, value=dados["nome_uf"].upper())
    sheet.cell(row=next_row, column=4, value=dados["codigo_do_municipio"])
    # Converte o valor para numérico, se for possível
    try:
        populacao_float = float(dados["populacao_ultimo_censo"])  # Use float() se for um número com casas decimais
    except ValueError:
        populacao_float = 0  # Ou outra ação adequada se a conversão falhar

    # Salva o valor numérico na célula
    sheet.cell(row=next_row, column=5, value=populacao_float)
    #sheet.cell(row=next_row, column=5, value=dados["populacao_ultimo_censo"])
    sheet.cell(row=next_row, column=6, value=dados["salario_medio_mensal"])
    sheet.cell(row=next_row, column=7, value=dados["matriculas_ensino_fundamental"])
    sheet.cell(row=next_row, column=8, value=dados["pip_per_capta"])
    sheet.cell(row=next_row, column=9, value=dados["mortalidade_infantil"])
    sheet.cell(row=next_row, column=10, value=dados["area_urbanizada"])
    sheet.cell(row=next_row, column=11, value=dados["area_unidade_territorial"])

    wb.save(caminho_arquivo)
    print(f"Dados do município {dados['nome_municipio']} salvos com sucesso na linha {next_row}.")
